# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 10:49:58 2024

@author: sophi
"""
#cd C:\Users\sophi\myCloud\Sophia\Thesis\Model\Jucar_model\Adrià\App_Jucar
#streamlit run Jucar_river_basin.py
# -*- coding: utf-8 -*-
"""
Model to run Vensim 
"""
### 0.Import functions and libraries
#work directory 
import os
os.chdir(r'C:\Users\sophi\myCloud\Sophia\Thesis\Model\Jucar_model\Adrià')
#import Python System Dynamics library to run Vensim
import pysd
#for interface
import dash
from dash import Dash, dcc, html, Input, Output, State, dash_table
import io
import base64
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math

### 1.Read data
#transform all the formulas into static number 
workbook = load_workbook('data_formulas.xlsx', data_only=True) 
sheet = workbook.active #it will read all the sheets 
#overwrite formulas with values
for row in sheet.iter_rows():
    for cell in row:
        if cell.data_type == 'f':  # if the cell contains a formula
            cell.value = cell.value  #replace the formula by the static value
workbook.save('data.xlsx')


# Import the data with all sheets (type = dictionary)
data_all = pd.read_excel("data.xlsx", sheet_name=None, engine="openpyxl")# List of sheet names
sheet_names = list(data_all.keys())
# Create the variable for each of the sheets (globals() allows you to modify the global namespace))
for sheet_name in sheet_names:
    globals()[f"data_{sheet_name}"] = pd.read_excel("data.xlsx", sheet_name=sheet_name,skiprows=1, engine="openpyxl")
    print(f"data_{sheet_name}")
    
### 2. Read Vensim Model
vensim_model = pysd.read_vensim('WEFE Jucar (Simple).mdl')    
variables_model = vensim_model.run(params={'INITIAL TIME': 1,'FINAL TIME': 120,'TIME STEP': 1})

DéfQecolAlar = variables_model["DéfQecolAlar"]


### 2.Test change QecolAlar (environmental flow downstream of Alarcon’s reservoir) = constant value

QecolAlar = data_Demandas["QecolAlar"]

